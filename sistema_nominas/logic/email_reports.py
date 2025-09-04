"""
Módulo para la generación de reportes Excel y TXT del proceso de envío de nóminas.
"""
import os
import pandas as pd
from datetime import datetime
from logic.formato_archivos import generar_nombre_archivo
from utils.logger import log_info, log_warning, log_error, log_debug


def generar_accion_requerida(status):
    """Genera sugerencias de acción según el tipo de error."""
    if "Sin NIF en PDF" in status:
        return "Verificar que el PDF contenga NIFs válidos"
    elif "NIF no encontrado en la lista" in status:
        return "Añadir empleado a la lista o verificar NIF"
    elif "Email inválido" in status:
        return "Corregir formato del email del empleado"
    elif "Sin datos" in status:
        return "Verificar datos completos del empleado"
    else:
        return "Revisar manualmente este empleado"


def generar_reporte_final(stats, todas_las_tareas_originales, config):
    """Genera un reporte final en Excel con TODOS los empleados y sus estados."""
    try:
        carpeta_mes = stats.get('carpeta_mes')
        if not carpeta_mes:
            log_warning("No se pudo generar reporte: falta carpeta_mes en stats")
            return
            
        fecha_actual = datetime.now()
        mes_nombre = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
            7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }[fecha_actual.month]
        timestamp = fecha_actual.strftime('%H%M%S')
        archivo_reporte = os.path.join(carpeta_mes, f"reporte_envio_{fecha_actual.year}_{mes_nombre}_{timestamp}.xlsx")
        
        log_info(f"Generando reporte con {len(todas_las_tareas_originales)} empleados totales")
        
        # Preparar datos para el reporte
        datos_reporte = []
        
        # Crear sets para búsqueda rápida
        enviados_exitosos = set()
        errores_dict = {}
        
        # Procesar lista de errores de envío
        for error_info in stats.get('errores_lista', []):
            key = f"{error_info['nombre']}|{error_info['email']}"
            errores_dict[key] = error_info['error']
        
        # Determinar qué se envió exitosamente (total - errores)
        total_enviados = stats.get('enviados', 0)
        if total_enviados > 0:
            # Los que se procesaron y enviaron exitosamente
            count_exitosos = 0
            for tarea in todas_las_tareas_originales:
                if tarea['status'] == '[OK]':  # Solo los que estaban OK para envío
                    key = f"{tarea['nombre']}|{tarea['email']}"
                    if key not in errores_dict and count_exitosos < total_enviados:
                        enviados_exitosos.add(key)
                        count_exitosos += 1
        
        # Procesar TODAS las tareas (incluyendo las que tenían errores en Paso 2)
        for tarea in todas_las_tareas_originales:
            key = f"{tarea['nombre']}|{tarea['email']}"
            
            # Determinar estado final según la lógica correcta
            if tarea['status'] != '[OK]':
                # Las que tenían errores en Paso 2 = PENDIENTE
                estado_envio = "PENDIENTE"
                observaciones = f"No procesado: {tarea['status']}"
            elif key in enviados_exitosos:
                # Las que se procesaron y enviaron bien = ENVIADO
                estado_envio = "ENVIADO"
                observaciones = "Enviado correctamente"
            elif key in errores_dict:
                # Las que se procesaron pero fallaron en el envío = ERROR
                estado_envio = "ERROR"
                observaciones = errores_dict[key]
            else:
                # Fallback (no debería pasar)
                estado_envio = "PENDIENTE"
                observaciones = "No procesado"
            
            # Separar nombre y apellidos para el reporte
            nombre_solo = tarea['nombre']
            apellidos_solo = tarea.get('apellidos', '')
            
            # Generar nombre del archivo 
            plantilla_archivo = config.get('Formato', 'archivo_nomina', fallback='{nombre}_Nomina_{mes}_{año}.pdf')
            nombre_archivo = generar_nombre_archivo(plantilla_archivo, nombre_solo, apellidos_solo)
            
            # Obtener la posición original si existe
            posicion_original = tarea.get('posicion_original', 'N/A')
            
            # Obtener fecha de procesamiento real o usar fecha por defecto
            fecha_procesamiento = tarea.get('fecha_procesamiento', 'No procesado')
            if fecha_procesamiento == 'No procesado' and estado_envio == 'PENDIENTE':
                fecha_procesamiento = 'No procesado'
            elif fecha_procesamiento == 'No procesado':
                fecha_procesamiento = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
            datos_reporte.append({
                'POS.': posicion_original,  # Primera columna: posición original
                'Página PDF': tarea['pagina'],
                'D.N.I.': tarea['nif'], 
                'Nombre': nombre_solo,
                'Apellidos': apellidos_solo,
                'Email': tarea['email'],
                'Archivo PDF': nombre_archivo,
                'Estado Envío': estado_envio,
                'Observaciones': observaciones,  # Mantener para lógica empresarial
                'Fecha Procesado': fecha_procesamiento
            })
        
        # Ordenar datos por posición original (de 0 al infinito)
        def obtener_posicion_para_ordenar(item):
            pos = item['POS.']
            # Si es 'N/A' o vacío, ponerlo al final
            if pos == 'N/A' or pos is None or pos == '':
                return float('inf')
            # Intentar convertir a número (puede venir como string desde Excel)
            try:
                return float(pos)
            except (ValueError, TypeError):
                return float('inf')
        
        datos_reporte.sort(key=obtener_posicion_para_ordenar)
        
        # DEBUG: Mostrar los primeros elementos para verificar ordenamiento
        log_info(f"[DEBUG] Primeros 5 elementos después de ordenar por posición:")
        for i, item in enumerate(datos_reporte[:5]):
            pos_raw = item['POS.']
            pos_num = obtener_posicion_para_ordenar(item)
            log_info(f"[DEBUG]   {i+1}. POS: '{pos_raw}' -> {pos_num}, Página: {item['Página PDF']}, Nombre: {item['Nombre']}")
        
        # Crear Excel con múltiples hojas
        _crear_excel_completo(archivo_reporte, datos_reporte, stats, config, todas_las_tareas_originales)
        
        # Crear TXT simple para referencia rápida
        _crear_reporte_txt(carpeta_mes, stats, archivo_reporte)
        
        # Actualizar stats con rutas de reportes
        stats['archivo_reporte_excel'] = archivo_reporte
        stats['archivo_resumen_txt'] = os.path.join(carpeta_mes, "resumen_proceso.txt")
        
    except Exception as e:
        log_error(f"[ERROR] Error generando reporte final: {e}")
        log_debug("Stack trace del error de reporte:", exc_info=True)


def _crear_excel_completo(archivo_reporte, datos_reporte, stats, config, todas_las_tareas_originales):
    """Crea el archivo Excel con todas las hojas y formato."""
    # Crear DataFrame y guardar Excel
    df = pd.DataFrame(datos_reporte)
    
    # Crear Excel con múltiples hojas y formato
    with pd.ExcelWriter(archivo_reporte, engine='openpyxl') as writer:
        # Hoja 1: Detalle completo
        df.to_excel(writer, sheet_name='Detalle Envios', index=False)
        worksheet1 = writer.sheets['Detalle Envios']
        
        # Aplicar formato a la hoja de detalles
        _aplicar_formato_detalle(worksheet1, datos_reporte)
        
        # Hoja 2: Resumen estadístico
        _crear_hoja_resumen(writer, stats, config)
        
        # Hoja 3: Empleados Pendientes
        _crear_hoja_pendientes(writer, todas_las_tareas_originales)
    
    log_info(f"Reporte generado: {archivo_reporte}")


def _aplicar_formato_detalle(worksheet, datos_reporte):
    """Aplica formato a la hoja de detalle de envíos."""
    # Importar estilos de openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Protection
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    # Definir colores
    color_success = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
    color_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro  
    color_pending = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo claro
    color_manual = PatternFill(start_color="B7E3FF", end_color="B7E3FF", fill_type="solid")   # Azul claro
    color_cabecera = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Gris claro
    font_cabecera = Font(bold=True)
    
    # Colorear cabeceras (fila 1)
    for cell in worksheet[1]:
        cell.fill = color_cabecera
        cell.font = font_cabecera
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Procesar todas las columnas para ajustar ancho
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for row_num, cell in enumerate(column, 1):
            try:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Ajustar ancho de columna con un mínimo y máximo
        adjusted_width = min(max(max_length + 2, 12), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Solo aplicar validación y formato si hay datos
    if len(datos_reporte) > 0:
        _aplicar_validacion_estado(worksheet, datos_reporte)
        _aplicar_formato_condicional(worksheet, datos_reporte, color_success, color_error, color_pending, color_manual)
        _aplicar_proteccion_hoja(worksheet, datos_reporte)


def _aplicar_validacion_estado(worksheet, datos_reporte):
    """Aplica validación dropdown para la columna Estado Envío."""
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Crear validación con las 4 opciones
    dv = DataValidation(
        type="list",
        formula1='"ENVIADO,ERROR,PENDIENTE,MANUAL"',
        allow_blank=False
    )
    dv.error = 'El valor debe ser: ENVIADO, ERROR, PENDIENTE o MANUAL'
    dv.errorTitle = 'Valor inválido'
    dv.prompt = 'Seleccione el estado del envío'
    dv.promptTitle = 'Estado del Envío'
    
    # Aplicar validación a todas las celdas de la columna H (excepto cabecera)
    rango_validacion = f"H2:H{len(datos_reporte) + 1}"
    dv.add(rango_validacion)
    worksheet.add_data_validation(dv)
    
    log_info(f"[OK] Dropdown agregado a rango: {rango_validacion}")


def _aplicar_formato_condicional(worksheet, datos_reporte, color_success, color_error, color_pending, color_manual):
    """Aplica formato condicional para colorear filas según el estado."""
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    
    # Rango completo de datos (todas las columnas, todas las filas con datos)  
    rango_datos = f"A2:J{len(datos_reporte) + 1}"
    
    # Regla 1: ENVIADO = Verde
    regla_enviado = Rule(
        type="expression",
        formula=[f'$H2="ENVIADO"'],  # Si columna H = "ENVIADO"
        dxf=DifferentialStyle(fill=color_success)
    )
    worksheet.conditional_formatting.add(rango_datos, regla_enviado)
    
    # Regla 2: ERROR = Rojo  
    regla_error = Rule(
        type="expression", 
        formula=[f'$H2="ERROR"'],  # Si columna H = "ERROR"
        dxf=DifferentialStyle(fill=color_error)
    )
    worksheet.conditional_formatting.add(rango_datos, regla_error)
    
    # Regla 3: PENDIENTE = Amarillo
    regla_pendiente = Rule(
        type="expression",
        formula=[f'$H2="PENDIENTE"'],  # Si columna H = "PENDIENTE" 
        dxf=DifferentialStyle(fill=color_pending)
    )
    worksheet.conditional_formatting.add(rango_datos, regla_pendiente)
    
    # Regla 4: MANUAL = Azul claro
    regla_manual = Rule(
        type="expression",
        formula=[f'$H2="MANUAL"'],  # Si columna H = "MANUAL"
        dxf=DifferentialStyle(fill=color_manual)
    )
    worksheet.conditional_formatting.add(rango_datos, regla_manual)
    
    log_info(f"[OK] Formato condicional agregado a rango: {rango_datos}")


def _aplicar_proteccion_hoja(worksheet, datos_reporte):
    """Aplica protección dejando editable solo la columna Estado Envío."""
    from openpyxl.styles import Protection
    
    # 1. Proteger toda la hoja
    for row in worksheet.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True, hidden=False)
    
    # 2. Desproteger solo la columna H "Estado Envío" (excepto cabecera)
    for row_num in range(2, len(datos_reporte) + 2):  # Desde fila 2 hasta la última con datos
        worksheet.cell(row=row_num, column=8).protection = Protection(locked=False, hidden=False)  # Columna H
    
    # 3. Activar protección de la hoja (sin contraseña para facilidad de uso)
    worksheet.protection.sheet = True
    worksheet.protection.enable()
    
    log_info("[OK] Protección aplicada: Solo columna 'Estado Envío' es editable")
    
    # Agregar lógica empresarial: Auto-llenar observaciones
    # Columna I = Observaciones, será oculta pero con fórmulas
    for row_num in range(2, len(datos_reporte) + 2):
        # Fórmula: Si H="MANUAL" → "Procesado manualmente", sino mantener observación original
        formula = f'=IF(H{row_num}="MANUAL","Procesado manualmente",I{row_num})'
        worksheet.cell(row=row_num, column=9).value = formula
        # Desproteger también columna I para que la fórmula funcione
        worksheet.cell(row=row_num, column=9).protection = Protection(locked=False, hidden=True)
    
    # Ocultar columna I "Observaciones" pero mantener funcionalidad
    worksheet.column_dimensions['I'].hidden = True
    
    log_info("[OK] Lógica empresarial agregada: Auto-llenado de observaciones")


def _crear_hoja_resumen(writer, stats, config):
    """Crea la hoja de resumen estadístico."""
    from openpyxl.styles import Alignment, PatternFill, Font
    
    resumen_data = {
        'Metrica': [
            'Total Procesadas',
            'Enviadas Exitosamente', 
            'Con Errores',
            'Tasa de Exito',
            'Email Remitente',
            'Fecha del Proceso',
            'Carpeta PDFs'
        ],
        'Valor': [
            stats['total'],
            stats['enviados'],
            stats['errores'],
            f"{(stats['enviados'] / stats['total'] * 100):.1f}%" if stats['total'] > 0 else "0%",
            config.get('Email', 'email_origen', fallback='N/A'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            stats.get('carpeta_pdfs', 'N/A')
        ]
    }
    df_resumen = pd.DataFrame(resumen_data)
    df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
    worksheet2 = writer.sheets['Resumen']
    
    # Definir colores para el resumen
    color_success = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    color_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    color_cabecera = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    font_cabecera = Font(bold=True)
    
    # Colorear cabeceras del resumen
    for cell in worksheet2[1]:
        cell.fill = color_cabecera
        cell.font = font_cabecera
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formato y colores para métricas
    for column in worksheet2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for row_num, cell in enumerate(column, 1):
            try:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Colorear métricas importantes
                if row_num > 1 and column_letter == 'A':  # Columna de métricas
                    if "Enviadas Exitosamente" in str(cell.value):
                        worksheet2.cell(row=row_num, column=1).fill = color_success
                        worksheet2.cell(row=row_num, column=2).fill = color_success
                    elif "Con Errores" in str(cell.value):
                        worksheet2.cell(row=row_num, column=1).fill = color_error  
                        worksheet2.cell(row=row_num, column=2).fill = color_error
                
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 15), 50)
        worksheet2.column_dimensions[column_letter].width = adjusted_width


def _crear_hoja_pendientes(writer, todas_las_tareas_originales):
    """Crea la hoja de empleados pendientes."""
    from openpyxl.styles import Alignment, PatternFill, Font
    
    # Hoja 3: Empleados Pendientes (los que NO se pudieron procesar)
    empleados_pendientes = [t for t in todas_las_tareas_originales if t['status'] != '[OK]']
    
    if empleados_pendientes:
        pendientes_data = []
        for tarea in empleados_pendientes:
            posicion_original = tarea.get('posicion_original', 'N/A')
            pendientes_data.append({
                'POS.': posicion_original,  # Primera columna: posición original
                'Página PDF': tarea['pagina'],
                'D.N.I.': tarea['nif'],
                'Nombre': tarea['nombre'],
                'Apellidos': tarea.get('apellidos', ''),
                'Email': tarea['email'],
                'Motivo Pendiente': tarea['status'].replace('[OK]', '').replace('[ERROR]', '').replace('[ADVERTENCIA]', '').strip(),
                'Acción Requerida': generar_accion_requerida(tarea['status'])
            })
        
        # Ordenar los datos de pendientes por posición también
        def obtener_posicion_pendientes(item):
            pos = item['POS.']
            # Si es 'N/A' o vacío, ponerlo al final
            if pos == 'N/A' or pos is None or pos == '':
                return float('inf')
            # Intentar convertir a número (puede venir como string desde Excel)
            try:
                return float(pos)
            except (ValueError, TypeError):
                return float('inf')
        
        pendientes_data.sort(key=obtener_posicion_pendientes)
        
        df_pendientes = pd.DataFrame(pendientes_data)
        df_pendientes.to_excel(writer, sheet_name='Pendientes', index=False)
        worksheet3 = writer.sheets['Pendientes']
        
        # Definir colores
        color_pending = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        color_cabecera = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        font_cabecera = Font(bold=True)
        
        # Colorear cabeceras de la hoja Pendientes
        for cell in worksheet3[1]:
            cell.fill = color_cabecera
            cell.font = font_cabecera
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar columnas y aplicar formato
        for column in worksheet3.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for row_num, cell in enumerate(column, 1):
                try:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Colorear todas las filas de datos con amarillo claro (pendiente)
                    if row_num > 1:  # Excepto cabecera
                        cell.fill = color_pending
                    
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 15), 50)
            worksheet3.column_dimensions[column_letter].width = adjusted_width
        
        log_info(f"[OK] Hoja 'Pendientes' agregada con {len(empleados_pendientes)} empleados")
    else:
        log_info("[OK] No hay empleados pendientes - todos se procesaron correctamente")


def _crear_reporte_txt(carpeta_mes, stats, archivo_reporte):
    """Crea un resumen simple en formato TXT."""
    archivo_txt = os.path.join(carpeta_mes, "resumen_proceso.txt")
    with open(archivo_txt, 'w', encoding='utf-8') as f:
        f.write("=" * 50 + "\n")
        f.write("RESUMEN DEL PROCESO DE ENVIO DE NOMINAS\n")
        f.write("=" * 50 + "\n")
        f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total procesadas: {stats['total']}\n")
        f.write(f"Enviadas exitosamente: {stats['enviados']}\n")
        f.write(f"Con errores: {stats['errores']}\n")
        f.write(f"Tasa de exito: {(stats['enviados'] / stats['total'] * 100):.1f}%\n" if stats['total'] > 0 else "Tasa de exito: 0%\n")
        f.write(f"\nCarpeta PDFs: {stats.get('carpeta_pdfs', 'N/A')}\n")
        f.write(f"Reporte detallado: {os.path.basename(archivo_reporte)}\n")
        
        if stats.get('errores_lista'):
            f.write(f"\nErrores encontrados ({len(stats['errores_lista'])}):\n")
            for error_info in stats['errores_lista']:
                f.write(f"- {error_info['nombre']} ({error_info['email']}): {error_info['error']}\n")
    
    log_info(f"Resumen TXT generado: {archivo_txt}")