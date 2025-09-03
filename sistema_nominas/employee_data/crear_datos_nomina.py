#!/usr/bin/env python3
"""
Convertir lista_empleados.csv al formato XLS de tu madre
"""

import csv
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Función para generar números de seguridad social falsos
def generar_nss():
    provincia = random.choice(['01', '02', '03', '04', '05', '06', '07', '08', '09', '10'])
    numero = str(random.randint(100000000, 999999999)).zfill(9)
    digito = str(random.randint(10, 99))
    return f"{provincia} {numero[:2]}{digito}"

# Leer CSV existente
empleados = []
csv_path = '/home/jon.garmilla.dev/Dev-Lab/Pdf-Manipulation-Test/Nominas_Dinanlag_Group/sistema_nominas/employee_data/lista_empleados.csv'

with open(csv_path, 'r', encoding='utf-8') as file:
    reader = csv.DictReader(file)
    for row in reader:
        empleados.append(row)

# Crear nuevo workbook
wb = Workbook()
ws = wb.active
ws.title = "DATOS PARA NÓMINA"

# Escribir encabezados con formato
headers = ['POS.', 'NOMBRE', 'APELLIDOS', 'D.N.I.', 'Nº SEGURIDAD SOCIAL', 'DIRECCIÓN DE CORREO ELECTRÓNICO']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Escribir datos de empleados
for idx, empleado in enumerate(empleados, 1):
    nombre_completo = empleado['NOMBRE']
    
    # Separar nombre y apellidos
    if ' ' in nombre_completo:
        partes = nombre_completo.split(' ', 1)
        nombre = partes[0]
        apellidos = partes[1] if len(partes) > 1 else ''
    else:
        nombre = nombre_completo
        apellidos = ''
    
    # Escribir fila (idx es 1-50, pero row es idx+1 porque fila 1 son headers)
    row = idx + 1
    ws.cell(row=row, column=1, value=idx-1)  # POS. (0-49)
    ws.cell(row=row, column=2, value=nombre)
    ws.cell(row=row, column=3, value=apellidos)
    ws.cell(row=row, column=4, value=empleado['NIF'])
    ws.cell(row=row, column=5, value=generar_nss())
    ws.cell(row=row, column=6, value=empleado['EMAIL'])

# Ajustar ancho de columnas
column_widths = [8, 15, 20, 12, 18, 35]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + col)].width = width

# Guardar archivo
output_path = '/home/jon.garmilla.dev/Dev-Lab/Pdf-Manipulation-Test/Nominas_Dinanlag_Group/sistema_nominas/employee_data/DATOS_PARA_NOMINA.xlsx'
wb.save(output_path)

print(f"✅ Archivo XLS creado: {output_path}")
print(f"📊 Total empleados: {len(empleados)}")
print(f"📋 Formato: {', '.join(headers)}")

# Mostrar muestra
print("\n=== MUESTRA (Primeros 3 empleados) ===")
for i in range(min(3, len(empleados))):
    empleado = empleados[i]
    nombre_completo = empleado['NOMBRE']
    if ' ' in nombre_completo:
        partes = nombre_completo.split(' ', 1)
        nombre = partes[0]
        apellidos = partes[1]
    else:
        nombre = nombre_completo
        apellidos = ''
    
    print(f"{i}: {nombre} | {apellidos} | {empleado['NIF']} | {empleado['EMAIL']}")