import os
import smtplib
import logging
import time
import socket
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pikepdf
import fitz  # PyMuPDF
import pandas as pd
from .formato_archivos import generar_nombre_archivo


def validar_email_basico(email):
    """Validación básica de formato de email."""
    if not email or not isinstance(email, str):
        return False
    
    email = email.strip().lower()
    
    # Verificaciones básicas
    if '@' not in email or '.' not in email:
        return False
        
    # Verificar que no empiece o termine con caracteres especiales
    if email.startswith(('@', '.', '-', '_')) or email.endswith(('.', '-', '_')):
        return False
        
    # Verificar longitud razonable
    if len(email) < 5 or len(email) > 100:
        return False
        
    # Verificar caracteres válidos (básico)
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def generar_estadisticas_envio(tareas):
    """Genera estadísticas detalladas de las tareas a procesar."""
    stats = {
        'total_validadas': len([t for t in tareas if t['status'] == '✅ OK']),
        'total_errores': len([t for t in tareas if t['status'].startswith('❌')]),
        'total_warnings': len([t for t in tareas if t['status'].startswith('⚠️')]),
        'emails_invalidos': 0,
        'emails_duplicados': 0
    }
    
    # Analizar emails
    emails_vistos = set()
    for tarea in tareas:
        if tarea['status'] == '✅ OK':
            email = tarea.get('email', '').strip().lower()
            
            if not validar_email_basico(email):
                stats['emails_invalidos'] += 1
            
            if email in emails_vistos:
                stats['emails_duplicados'] += 1
            else:
                emails_vistos.add(email)
    
    return stats


class RobustEmailSender:
    """Cliente SMTP robusto con reintentos, timeouts y manejo de errores."""
    
    def __init__(self, config):
        self.config = config
        self.servidor_smtp = config.get('SMTP', 'servidor', fallback='smtp.gmail.com')
        self.puerto_smtp = int(config.get('SMTP', 'puerto', fallback='587'))
        self.timeout = int(config.get('SMTP', 'timeout', fallback='30'))
        self.max_reintentos = int(config.get('SMTP', 'max_reintentos', fallback='3'))
        self.delay_entre_emails = float(config.get('SMTP', 'delay_segundos', fallback='1.0'))
        self.server = None
        self.conexiones_fallidas = 0
        
    def conectar(self, email_origen, password):
        """Establece conexión SMTP con reintentos."""
        for intento in range(self.max_reintentos):
            try:
                logging.info(f"Conectando a {self.servidor_smtp}:{self.puerto_smtp} (intento {intento + 1})")
                
                # Configurar timeout a nivel socket
                socket.setdefaulttimeout(self.timeout)
                
                self.server = smtplib.SMTP(self.servidor_smtp, self.puerto_smtp, timeout=self.timeout)
                self.server.set_debuglevel(0)  # 0=sin debug, 1=debug básico, 2=debug completo
                
                logging.info("Iniciando TLS...")
                self.server.starttls()
                
                logging.info("Autenticando...")
                self.server.login(email_origen, password)
                
                logging.info("✅ Conexión SMTP establecida exitosamente")
                self.conexiones_fallidas = 0
                return True
                
            except smtplib.SMTPAuthenticationError as e:
                logging.error(f"❌ Error de autenticación: {e}")
                logging.error("Verificar credenciales de email y contraseñas de aplicación")
                return False
                
            except smtplib.SMTPServerDisconnected as e:
                logging.warning(f"⚠️  Servidor desconectado (intento {intento + 1}): {e}")
                
            except socket.timeout as e:
                logging.warning(f"⚠️  Timeout de conexión (intento {intento + 1}): {e}")
                
            except socket.gaierror as e:
                logging.error(f"❌ Error de DNS/red: {e}")
                return False
                
            except Exception as e:
                logging.warning(f"⚠️  Error de conexión (intento {intento + 1}): {type(e).__name__}: {e}")
            
            # Backoff exponencial entre reintentos
            if intento < self.max_reintentos - 1:
                delay = (2 ** intento) * 2  # 2, 4, 8 segundos
                logging.info(f"Esperando {delay}s antes del siguiente intento...")
                time.sleep(delay)
        
        logging.error(f"❌ Failed to connect after {self.max_reintentos} attempts")
        self.conexiones_fallidas += 1
        return False
    
    def enviar_email(self, msg, email_destino, max_reintentos=None):
        """Envía un email con reintentos automáticos."""
        if max_reintentos is None:
            max_reintentos = self.max_reintentos
            
        for intento in range(max_reintentos):
            try:
                # Verificar conexión
                if not self.server:
                    raise smtplib.SMTPServerDisconnected("No hay conexión activa")
                
                # Enviar mensaje
                self.server.send_message(msg)
                
                # Rate limiting - pausa entre emails
                if self.delay_entre_emails > 0:
                    time.sleep(self.delay_entre_emails)
                    
                return True
                
            except smtplib.SMTPServerDisconnected as e:
                logging.warning(f"⚠️  Servidor desconectado durante envío (intento {intento + 1}): {e}")
                # Intentar reconectar
                email_origen = self.config.get('Email', 'email_origen')
                password = self.config.get('Email', 'password')
                if self.conectar(email_origen, password):
                    continue  # Reintentar envío
                else:
                    return False
                    
            except smtplib.SMTPRecipientsRefused as e:
                logging.error(f"❌ Email rechazado por el servidor: {email_destino} - {e}")
                return False  # No reintentar, email inválido
                
            except smtplib.SMTPDataError as e:
                logging.error(f"❌ Error de datos SMTP: {e}")
                return False  # No reintentar, problema con el mensaje
                
            except socket.timeout as e:
                logging.warning(f"⚠️  Timeout enviando email (intento {intento + 1}): {e}")
                
            except Exception as e:
                logging.warning(f"⚠️  Error enviando email (intento {intento + 1}): {type(e).__name__}: {e}")
            
            # Pausa antes del siguiente intento
            if intento < max_reintentos - 1:
                delay = min((2 ** intento), 10)  # Max 10 segundos
                logging.info(f"Esperando {delay}s antes de reintentar envío...")
                time.sleep(delay)
        
        logging.error(f"❌ Failed to send email to {email_destino} after {max_reintentos} attempts")
        return False
    
    def cerrar(self):
        """Cierra la conexión SMTP de forma segura."""
        if self.server:
            try:
                self.server.quit()
                logging.info("✅ Conexión SMTP cerrada correctamente")
            except:
                try:
                    self.server.close()
                except:
                    pass
            finally:
                self.server = None


def generar_accion_requerida(status):
    """Genera sugerencias de acción según el tipo de error."""
    if "Sin NIF en PDF" in status:
        return "Verificar que el PDF contenga NIFs válidos"
    elif "NIF no encontrado en la lista" in status:
        return "Añadir empleado a la lista o verificar NIF"
    elif "Email inválido" in status:
        return "Corregir formato del email del empleado"
    elif "Sin datos" in status:
        return "Verificar datos completos del empleado"
    else:
        return "Revisar manualmente este empleado"


def generar_reporte_final(stats, todas_las_tareas_originales, config):
    """Genera un reporte final en Excel con TODOS los empleados y sus estados."""
    try:
        carpeta_mes = stats.get('carpeta_mes')
        if not carpeta_mes:
            logging.warning("No se pudo generar reporte: falta carpeta_mes en stats")
            return
            
        fecha_actual = datetime.now()
        mes_nombre = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
            7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }[fecha_actual.month]
        timestamp = fecha_actual.strftime('%H%M%S')
        archivo_reporte = os.path.join(carpeta_mes, f"reporte_envio_{fecha_actual.year}_{mes_nombre}_{timestamp}.xlsx")
        
        logging.info(f"📊 Generando reporte con {len(todas_las_tareas_originales)} empleados totales")
        
        # Preparar datos para el reporte
        datos_reporte = []
        
        # Crear sets para búsqueda rápida
        enviados_exitosos = set()
        errores_dict = {}
        
        # Procesar lista de errores de envío
        for error_info in stats.get('errores_lista', []):
            key = f"{error_info['nombre']}|{error_info['email']}"
            errores_dict[key] = error_info['error']
        
        # Determinar qué se envió exitosamente (total - errores)
        total_enviados = stats.get('enviados', 0)
        if total_enviados > 0:
            # Los que se procesaron y enviaron exitosamente
            count_exitosos = 0
            for tarea in todas_las_tareas_originales:
                if tarea['status'] == '✅ OK':  # Solo los que estaban OK para envío
                    key = f"{tarea['nombre']}|{tarea['email']}"
                    if key not in errores_dict and count_exitosos < total_enviados:
                        enviados_exitosos.add(key)
                        count_exitosos += 1
        
        # Procesar TODAS las tareas (incluyendo las que tenían errores en Paso 2)
        for tarea in todas_las_tareas_originales:
            key = f"{tarea['nombre']}|{tarea['email']}"
            
            # Determinar estado final según la lógica correcta
            if tarea['status'] != '✅ OK':
                # Las que tenían errores en Paso 2 = PENDIENTE
                estado_envio = "PENDIENTE"
                observaciones = f"No procesado: {tarea['status']}"
            elif key in enviados_exitosos:
                # Las que se procesaron y enviaron bien = ENVIADO
                estado_envio = "ENVIADO"
                observaciones = "Enviado correctamente"
            elif key in errores_dict:
                # Las que se procesaron pero fallaron en el envío = ERROR
                estado_envio = "ERROR"
                observaciones = errores_dict[key]
            else:
                # Fallback (no debería pasar)
                estado_envio = "PENDIENTE"
                observaciones = "No procesado"
            
            # Separar nombre y apellidos para el reporte
            nombre_solo = tarea['nombre']
            apellidos_solo = tarea.get('apellidos', '')
            
            # Generar nombre del archivo 
            plantilla_archivo = config.get('Formato', 'archivo_nomina', fallback='{nombre}_Nomina_{mes}_{año}.pdf')
            nombre_archivo = generar_nombre_archivo(plantilla_archivo, nombre_solo, apellidos_solo)
            
            datos_reporte.append({
                'Página PDF': tarea['pagina'],
                'D.N.I.': tarea['nif'], 
                'Nombre': nombre_solo,
                'Apellidos': apellidos_solo,
                'Email': tarea['email'],
                'Archivo PDF': nombre_archivo,
                'Estado Envío': estado_envio,
                'Observaciones': observaciones,  # Mantener para lógica empresarial
                'Fecha Procesado': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Crear DataFrame y guardar Excel
        df = pd.DataFrame(datos_reporte)
        
        # Crear Excel con múltiples hojas y formato
        with pd.ExcelWriter(archivo_reporte, engine='openpyxl') as writer:
            # Hoja 1: Detalle completo
            df.to_excel(writer, sheet_name='Detalle Envios', index=False)
            worksheet1 = writer.sheets['Detalle Envios']
            
            # Centrar todas las celdas y ajustar ancho de columnas con colores
            from openpyxl.styles import Alignment, PatternFill, Font, Protection
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle
            
            # Definir colores
            color_success = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
            color_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro  
            color_pending = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo claro
            color_cabecera = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Gris claro
            font_cabecera = Font(bold=True)
            
            # Colorear cabeceras (fila 1)
            for cell in worksheet1[1]:
                cell.fill = color_cabecera
                cell.font = font_cabecera
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Procesar todas las columnas para ajustar ancho
            for column in worksheet1.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for row_num, cell in enumerate(column, 1):
                    try:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # El coloreo ahora se hará con formato condicional al final
                        
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Ajustar ancho de columna con un mínimo y máximo
                adjusted_width = min(max(max_length + 2, 12), 50)
                worksheet1.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar dropdown de validación para columna "Estado Envío" (columna F)
            if len(datos_reporte) > 0:  # Solo si hay datos
                # Crear validación con las 3 opciones
                dv = DataValidation(
                    type="list",
                    formula1='"ENVIADO,ERROR,PENDIENTE"',
                    allow_blank=False
                )
                dv.error = 'El valor debe ser: ENVIADO, ERROR o PENDIENTE'
                dv.errorTitle = 'Valor inválido'
                dv.prompt = 'Seleccione el estado del envío'
                dv.promptTitle = 'Estado del Envío'
                
                # Aplicar validación a todas las celdas de la columna G (excepto cabecera)
                rango_validacion = f"G2:G{len(datos_reporte) + 1}"
                dv.add(rango_validacion)
                worksheet1.add_data_validation(dv)
                
                logging.info(f"✅ Dropdown agregado a rango: {rango_validacion}")
                
                # Agregar formato condicional para coloreo dinámico de filas
                # Rango completo de datos (todas las columnas, todas las filas con datos)  
                rango_datos = f"A2:I{len(datos_reporte) + 1}"
                
                # Regla 1: ENVIADO = Verde
                regla_enviado = Rule(
                    type="expression",
                    formula=[f'$G2="ENVIADO"'],  # Si columna G = "ENVIADO"
                    dxf=DifferentialStyle(fill=color_success)
                )
                worksheet1.conditional_formatting.add(rango_datos, regla_enviado)
                
                # Regla 2: ERROR = Rojo  
                regla_error = Rule(
                    type="expression", 
                    formula=[f'$G2="ERROR"'],  # Si columna G = "ERROR"
                    dxf=DifferentialStyle(fill=color_error)
                )
                worksheet1.conditional_formatting.add(rango_datos, regla_error)
                
                # Regla 3: PENDIENTE = Amarillo
                regla_pendiente = Rule(
                    type="expression",
                    formula=[f'$G2="PENDIENTE"'],  # Si columna G = "PENDIENTE" 
                    dxf=DifferentialStyle(fill=color_pending)
                )
                worksheet1.conditional_formatting.add(rango_datos, regla_pendiente)
                
                logging.info(f"✅ Formato condicional agregado a rango: {rango_datos}")
                
                # Protección de celdas: Solo permitir editar columna "Estado Envío"
                # 1. Proteger toda la hoja
                for row in worksheet1.iter_rows():
                    for cell in row:
                        cell.protection = Protection(locked=True, hidden=False)
                
                # 2. Desproteger solo la columna G "Estado Envío" (excepto cabecera)
                for row_num in range(2, len(datos_reporte) + 2):  # Desde fila 2 hasta la última con datos
                    worksheet1.cell(row=row_num, column=7).protection = Protection(locked=False, hidden=False)  # Columna G
                
                # 3. Activar protección de la hoja (sin contraseña para facilidad de uso)
                worksheet1.protection.sheet = True
                worksheet1.protection.enable()
                
                logging.info("🔒 Protección aplicada: Solo columna 'Estado Envío' es editable")
                
                # Agregar lógica empresarial: Auto-llenar observaciones
                # Columna H = Observaciones, será oculta pero con fórmulas
                for row_num in range(2, len(datos_reporte) + 2):
                    # Fórmula: Si G="ENVIADO" y H era "No procesado" → "Enviado manualmente"  
                    formula = f'=IF(AND(G{row_num}="ENVIADO",H{row_num}="No procesado"),"Enviado manualmente",H{row_num})'
                    worksheet1.cell(row=row_num, column=8).value = formula
                    # Desproteger también columna H para que la fórmula funcione
                    worksheet1.cell(row=row_num, column=8).protection = Protection(locked=False, hidden=True)
                
                # Ocultar columna H "Observaciones" pero mantener funcionalidad
                worksheet1.column_dimensions['H'].hidden = True
                
                logging.info("📝 Lógica empresarial agregada: Auto-llenado de observaciones")
            
            # Hoja 2: Resumen estadístico  
            resumen_data = {
                'Metrica': [
                    'Total Procesadas',
                    'Enviadas Exitosamente', 
                    'Con Errores',
                    'Tasa de Exito',
                    'Email Remitente',
                    'Fecha del Proceso',
                    'Carpeta PDFs'
                ],
                'Valor': [
                    stats['total'],
                    stats['enviados'],
                    stats['errores'],
                    f"{(stats['enviados'] / stats['total'] * 100):.1f}%" if stats['total'] > 0 else "0%",
                    config.get('Email', 'email_origen', fallback='N/A'),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    stats.get('carpeta_pdfs', 'N/A')
                ]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            worksheet2 = writer.sheets['Resumen']
            
            # Centrar y ajustar resumen también con colores
            # Colorear cabeceras del resumen
            for cell in worksheet2[1]:
                cell.fill = color_cabecera
                cell.font = font_cabecera
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet2.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for row_num, cell in enumerate(column, 1):
                    try:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Colorear métricas importantes
                        if row_num > 1 and column_letter == 'A':  # Columna de métricas
                            if "Enviadas Exitosamente" in str(cell.value):
                                worksheet2.cell(row=row_num, column=1).fill = color_success
                                worksheet2.cell(row=row_num, column=2).fill = color_success
                            elif "Con Errores" in str(cell.value):
                                worksheet2.cell(row=row_num, column=1).fill = color_error  
                                worksheet2.cell(row=row_num, column=2).fill = color_error
                        
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 15), 50)
                worksheet2.column_dimensions[column_letter].width = adjusted_width
            
            # Hoja 3: Empleados Pendientes (los que NO se pudieron procesar)
            empleados_pendientes = [t for t in todas_las_tareas_originales if t['status'] != '✅ OK']
            
            if empleados_pendientes:
                pendientes_data = []
                for tarea in empleados_pendientes:
                    pendientes_data.append({
                        'Página PDF': tarea['pagina'],
                        'D.N.I.': tarea['nif'],
                        'Nombre': tarea['nombre'],
                        'Apellidos': tarea.get('apellidos', ''),
                        'Email': tarea['email'],
                        'Motivo Pendiente': tarea['status'].replace('✅', '').replace('❌', '').replace('⚠️', '').strip(),
                        'Acción Requerida': generar_accion_requerida(tarea['status'])
                    })
                
                df_pendientes = pd.DataFrame(pendientes_data)
                df_pendientes.to_excel(writer, sheet_name='Pendientes', index=False)
                worksheet3 = writer.sheets['Pendientes']
                
                # Colorear cabeceras de la hoja Pendientes
                for cell in worksheet3[1]:
                    cell.fill = color_cabecera
                    cell.font = font_cabecera
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar columnas y aplicar formato
                for column in worksheet3.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for row_num, cell in enumerate(column, 1):
                        try:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Colorear todas las filas de datos con amarillo claro (pendiente)
                            if row_num > 1:  # Excepto cabecera
                                cell.fill = color_pending
                            
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max(max_length + 2, 15), 50)
                    worksheet3.column_dimensions[column_letter].width = adjusted_width
                
                logging.info(f"✅ Hoja 'Pendientes' agregada con {len(empleados_pendientes)} empleados")
            else:
                logging.info("✅ No hay empleados pendientes - todos se procesaron correctamente")
        
        logging.info(f"Reporte generado: {archivo_reporte}")
        
        # También crear un TXT simple para referencia rápida
        archivo_txt = os.path.join(carpeta_mes, "resumen_proceso.txt")
        with open(archivo_txt, 'w', encoding='utf-8') as f:
            f.write("=" * 50 + "\n")
            f.write("RESUMEN DEL PROCESO DE ENVIO DE NOMINAS\n")
            f.write("=" * 50 + "\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total procesadas: {stats['total']}\n")
            f.write(f"Enviadas exitosamente: {stats['enviados']}\n")
            f.write(f"Con errores: {stats['errores']}\n")
            f.write(f"Tasa de exito: {(stats['enviados'] / stats['total'] * 100):.1f}%\n" if stats['total'] > 0 else "Tasa de exito: 0%\n")
            f.write(f"\nCarpeta PDFs: {stats.get('carpeta_pdfs', 'N/A')}\n")
            f.write(f"Reporte detallado: {os.path.basename(archivo_reporte)}\n")
            
            if stats.get('errores_lista'):
                f.write(f"\nErrores encontrados ({len(stats['errores_lista'])}):\n")
                for error_info in stats['errores_lista']:
                    f.write(f"- {error_info['nombre']} ({error_info['email']}): {error_info['error']}\n")
        
        logging.info(f"Resumen TXT generado: {archivo_txt}")
        
        # Actualizar stats con rutas de reportes
        stats['archivo_reporte_excel'] = archivo_reporte
        stats['archivo_resumen_txt'] = archivo_txt
        
    except Exception as e:
        logging.error(f"❌ Error generando reporte final: {e}")
        logging.debug("🐛 Stack trace del error de reporte:", exc_info=True)


def enviar_nominas_worker(
    pdf_path, tareas, config, status_callback, progress_callback, stop_event=None
):
    """Worker que procesa y envía las nóminas en un hilo separado."""
    # Configurar logging detallado
    import os
    from datetime import datetime
    
    # Crear carpeta de logs si no existe
    logs_dir = 'logs'
    os.makedirs(logs_dir, exist_ok=True)
    
    # Archivo de log con timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(logs_dir, f'envio_nominas_{timestamp}.log')
    
    # Configurar logging con más detalle
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - [%(levelname)8s] - %(funcName)s:%(lineno)d - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ],
        force=True  # Sobrescribir configuración anterior
    )
    
    logging.info(f"📝 Sistema de logging iniciado. Archivo: {log_file}")
    logging.info(f"🖥️  Sistema: {os.name} - Python: {__import__('sys').version.split()[0]}")
    logging.info("🚀 Iniciando el proceso de envío de nóminas.")
    
    # Estadísticas de envío
    stats = {
        'total': 0,
        'enviados': 0, 
        'errores': 0,
        'saltados': 0,
        'errores_lista': []
    }
    
    email_sender = None
    try:
        email_origen = config.get('Email', 'email_origen')
        password = config.get('Email', 'password')
        
        logging.debug(f"🔑 Email configurado: {email_origen}")
        logging.debug(f"🔑 Password configurado: {'Sí' if password else 'No'}")
        
        if not email_origen or not password:
            logging.error("❌ Credenciales de email no configuradas")
            raise ValueError("❌ Credenciales de email no configuradas.")

        # Inicializar cliente robusto de email
        email_sender = RobustEmailSender(config)
        
        # Establecer conexión con reintentos automáticos
        logging.debug("🔗 Intentando establecer conexión SMTP...")
        if not email_sender.conectar(email_origen, password):
            logging.error("❌ Falló la conexión SMTP después de reintentos")
            raise ConnectionError("❌ No se pudo establecer conexión SMTP después de varios intentos")
        logging.info("✅ Conexión SMTP establecida correctamente")

        # Análisis previo de las tareas
        pre_stats = generar_estadisticas_envio(tareas)
        logging.info(f"📊 ANÁLISIS PREVIO:")
        logging.info(f"   ✅ Validadas para envío: {pre_stats['total_validadas']}")
        logging.info(f"   ❌ Con errores: {pre_stats['total_errores']}")
        logging.info(f"   ⚠️  Con warnings: {pre_stats['total_warnings']}")
        if pre_stats['emails_invalidos'] > 0:
            logging.warning(f"   📧 Emails con formato inválido: {pre_stats['emails_invalidos']}")
        if pre_stats['emails_duplicados'] > 0:
            logging.warning(f"   🔄 Emails duplicados detectados: {pre_stats['emails_duplicados']}")

        doc_maestro = fitz.open(pdf_path)
        tareas_a_enviar = [t for t in tareas if t['status'] == '✅ OK']
        # Ordenar por número de página para procesar de arriba hacia abajo
        tareas_a_enviar.sort(key=lambda x: x['pagina'])
        
        stats['total'] = len(tareas_a_enviar)
        logging.info(f"🚀 Iniciando procesamiento de {stats['total']} nóminas.")
        
        # Crear estructura organizada por mes/año
        base_dir = config.get('Carpetas', 'salida', fallback='nominas_individuales')
        fecha_actual = datetime.now()
        mes_nombre = fecha_actual.strftime('%B').lower()  # septiembre
        año = fecha_actual.year
        
        # Carpeta principal: nominas_2025_09/
        carpeta_mes = os.path.join(base_dir, f"nominas_{año}_{fecha_actual.month:02d}")
        
        # Subcarpeta para PDFs: nominas_2025_09/pdfs_individuales/
        output_dir = os.path.join(carpeta_mes, "pdfs_individuales")
        os.makedirs(output_dir, exist_ok=True)
        
        # Guardar carpeta para el reporte
        stats['carpeta_mes'] = carpeta_mes
        
        logging.info(f"📁 Estructura creada: {carpeta_mes}")
        logging.info(f"📁 PDFs se guardarán en: {output_dir}")
        
        # Copiar PDF original a la carpeta del mes
        try:
            import shutil
            nombre_pdf_original = os.path.basename(pdf_path)
            destino_pdf_original = os.path.join(carpeta_mes, nombre_pdf_original)
            shutil.copy2(pdf_path, destino_pdf_original)
            logging.info(f"📄 PDF original copiado: {destino_pdf_original}")
        except Exception as e:
            logging.error(f"❌ Error al copiar PDF original: {e}")
        
        # Guardar rutas para el reporte final
        stats['carpeta_mes'] = carpeta_mes
        stats['carpeta_pdfs'] = output_dir

        # Procesar cada nómina con recuperación de errores
        for i, tarea in enumerate(tareas_a_enviar):
            # Verificar si se debe cancelar el proceso
            if stop_event and stop_event.is_set():
                logging.info("🛑 Proceso de envío cancelado por el usuario.")
                status_callback("proceso_cancelado", "Proceso cancelado", "cancelled")
                break
                
            nombre = tarea['nombre']
            nif = tarea['nif']
            email_destino = tarea['email']
            
            logging.info(
                f"📧 Procesando {i+1}/{stats['total']}: "
                f"{nombre} → {email_destino}"
            )
            
            tarea_exitosa = False
            error_msg = ""
            
            try:
                status_callback(f"pagina_{tarea['pagina']}", "Procesando PDF...", "processing")
                
                # 1. Validar email antes de procesar
                if not validar_email_basico(email_destino):
                    raise ValueError(f"Formato de email inválido: {email_destino}")
                
                # 2. Extraer y guardar PDF individual
                doc_individual = fitz.open()
                doc_individual.insert_pdf(
                    doc_maestro, from_page=tarea['pagina'] - 1,
                    to_page=tarea['pagina'] - 1)
                
                temp_pdf_path = os.path.join(output_dir, f"temp_{nif}.pdf")
                doc_individual.save(temp_pdf_path)
                doc_individual.close()

                # 3. Generar nombre de archivo personalizado  
                plantilla_archivo = config.get('Formato', 'archivo_nomina', 
                                              fallback='{nombre}_Nomina_{mes}_{año}.pdf')
                
                # Intentar obtener apellido si existe en la tarea, sino separar del nombre
                apellido_empleado = tarea.get('apellido', '')
                if not apellido_empleado and ' ' in nombre:
                    # Si no hay apellido separado, intentar separar del nombre completo
                    partes = nombre.strip().split(' ', 1)
                    nombre_empleado = partes[0]
                    apellido_empleado = partes[1] if len(partes) > 1 else ''
                else:
                    nombre_empleado = nombre
                
                nombre_archivo = generar_nombre_archivo(plantilla_archivo, nombre_empleado, apellido_empleado)
                
                pdf_encriptado_path = os.path.join(output_dir, nombre_archivo)
                
                # Usar contraseña de autor si está configurada
                password_autor = config.get('PDF', 'password_autor', fallback='')
                owner_password = password_autor if password_autor else nif
                
                logging.debug(f"Encriptando PDF para {nombre}")
                with pikepdf.open(temp_pdf_path) as pdf:
                    pdf.save(
                        pdf_encriptado_path,
                        encryption=pikepdf.Encryption(owner=owner_password, user=nif, R=4)
                    )
                os.remove(temp_pdf_path)

                # 4. Preparar email
                status_callback(f"pagina_{tarea['pagina']}", "Enviando email...", "processing")
                
                msg = MIMEMultipart()
                msg['From'] = email_origen
                msg['To'] = email_destino
                msg['Subject'] = f"Tu nómina de {datetime.now().strftime('%B %Y')}"
                
                cuerpo = (
                    f"Hola {nombre},\n\n"
                    f"Adjuntamos tu nómina correspondiente a {datetime.now().strftime('%B %Y')}.\n"
                    f"La contraseña para abrir el archivo es tu NIF.\n\n"
                    f"Saludos cordiales."
                )
                msg.attach(MIMEText(cuerpo, 'plain'))

                # Adjuntar PDF encriptado
                with open(pdf_encriptado_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{os.path.basename(pdf_encriptado_path)}"'
                )
                msg.attach(part)
                
                # 5. Enviar con reintentos automáticos
                envio_exitoso = email_sender.enviar_email(msg, email_destino)
                if envio_exitoso:
                    # Éxito - MANTENER el PDF para archivo
                    status_callback(f"pagina_{tarea['pagina']}", "SUCCESS", "sent")
                    stats['enviados'] += 1
                    logging.info(f"Email enviado exitosamente a {email_destino} (total enviados: {stats['enviados']})")
                    tarea_exitosa = True
                else:
                    # Error en envío (ya reintentado automáticamente)
                    error_msg = f"Fallo en envío después de reintentos"
                    stats['errores'] += 1
                    # Mantener PDF para inspección manual
                    logging.error(f"Email fallido a {email_destino} (total errores: {stats['errores']})")
                    
                    # Agregar a lista de errores
                    stats['errores_lista'].append({
                        'nombre': nombre,
                        'email': email_destino,
                        'error': error_msg
                    })
                
            except Exception as e:
                # Error en procesamiento del PDF o preparación del email
                error_msg = f"{type(e).__name__}: {str(e)[:100]}"
                stats['errores'] += 1
                logging.error(f"❌ Error procesando {nombre}: {error_msg}")
                logging.debug(f"🐛 Stack trace completo:", exc_info=True)
                
                # Limpiar archivos temporales en caso de error
                try:
                    if 'temp_pdf_path' in locals() and os.path.exists(temp_pdf_path):
                        os.remove(temp_pdf_path)
                    if 'pdf_encriptado_path' in locals() and os.path.exists(pdf_encriptado_path):
                        os.remove(pdf_encriptado_path)
                except:
                    pass
            
            # Actualizar estado final
            if not tarea_exitosa:
                status_callback(f"pagina_{tarea['pagina']}", f"ERROR: {error_msg}", "error")
                # Si no se agregó ya a la lista de errores, agregarlo ahora
                if not any(err['email'] == email_destino and err['nombre'] == nombre for err in stats['errores_lista']):
                    stats['errores_lista'].append({
                        'nombre': nombre,
                        'email': email_destino,
                        'error': error_msg
                    })
            
            # Actualizar progreso
            progress_callback((i + 1) / stats['total'] * 100)

        # Cerrar conexión de forma segura
        if email_sender:
            email_sender.cerrar()
        
        # Cerrar documento PDF maestro
        if doc_maestro:
            doc_maestro.close()
        
        # Resumen final
        logging.info("=" * 50)
        logging.info("📊 RESUMEN DEL ENVÍO:")
        logging.info(f"   📋 Total procesadas: {stats['total']}")
        logging.info(f"   ✅ Enviadas exitosamente: {stats['enviados']}")
        logging.info(f"   ❌ Con errores: {stats['errores']}")
        if stats['errores'] > 0:
            logging.info("   💥 Errores encontrados:")
            for error_info in stats['errores_lista'][:5]:  # Mostrar solo primeros 5
                logging.info(f"      • {error_info['nombre']}: {error_info['error']}")
            if len(stats['errores_lista']) > 5:
                logging.info(f"      ... y {len(stats['errores_lista']) - 5} errores más")
        
        tasa_exito = (stats['enviados'] / stats['total'] * 100) if stats['total'] > 0 else 0
        logging.info(f"   📈 Tasa de éxito: {tasa_exito:.1f}%")
        logging.info("=" * 50)
        
        # Generar reporte final con TODAS las tareas originales (para mostrar PENDIENTES)
        generar_reporte_final(stats, tareas, config)
        
        # Pasar estadísticas finales al callback
        logging.info(f"📊 ENVIANDO ESTADÍSTICAS FINALES: enviados={stats['enviados']}, errores={stats['errores']}, total={stats['total']}")
        status_callback("estadisticas_finales", "", "completed", stats)
        
        if stats['enviados'] > 0:
            logging.info("🎉 Proceso de envío de nóminas completado.")
        else:
            logging.warning("⚠️  No se enviaron nóminas exitosamente.")
            
    except Exception as e:
        logging.error(f"❌ Error crítico en el proceso: {type(e).__name__}: {e}")
        logging.debug(f"🐛 Stack trace crítico:", exc_info=True)
        status_callback("error_general", f"Error crítico: {str(e)[:100]}", "error")
        
        # Intentar cerrar conexiones en caso de error
        try:
            if email_sender:
                email_sender.cerrar()
        except:
            pass
            
    finally:
        # Asegurar que se complete el progreso
        progress_callback(-1)
